#!/usr/bin/env python3

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
import sys
from typing import Any, Optional, Tuple
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


EVENT_FIELDS = [
    "datasetName",
    "eventID",
    "eventDate",
    "institutionCode",
    "decimalLongitude",
    "decimalLatitude",
    "countryCode",
    "geodeticDatum",
    "minimumDepthInMeters",
    "maximumDepthInMeters",
    "sampleSizeValue",
    "sampleSizeUnit",
    "samplingProtocol",
]

OCCURRENCE_FIELDS = [
    "eventID",
    "occurrenceID",
    "scientificName",
    "scientificNameID",
    "taxonRank",
    "kingdom",
    "basisOfRecord",
    "occurrenceStatus",
    "verbatimIdentification",
    "identifiedBy",
    "identificationVerificationStatus",
    "identificationReferences",
    "associatedMedia",
]

META_EVENT_ID_INDEX = 1
META_OCCURRENCE_CORE_ID_INDEX = 0

META_EVENT_TERMS = {
    "datasetName": "http://rs.tdwg.org/dwc/terms/datasetName",
    "eventID": "http://rs.tdwg.org/dwc/terms/eventID",
    "eventDate": "http://rs.tdwg.org/dwc/terms/eventDate",
    "institutionCode": "http://rs.tdwg.org/dwc/terms/institutionCode",
    "decimalLongitude": "http://rs.tdwg.org/dwc/terms/decimalLongitude",
    "decimalLatitude": "http://rs.tdwg.org/dwc/terms/decimalLatitude",
    "countryCode": "http://rs.tdwg.org/dwc/terms/countryCode",
    "geodeticDatum": "http://rs.tdwg.org/dwc/terms/geodeticDatum",
    "minimumDepthInMeters": "http://rs.tdwg.org/dwc/terms/minimumDepthInMeters",
    "maximumDepthInMeters": "http://rs.tdwg.org/dwc/terms/maximumDepthInMeters",
    "sampleSizeValue": "http://rs.tdwg.org/dwc/terms/sampleSizeValue",
    "sampleSizeUnit": "http://rs.tdwg.org/dwc/terms/sampleSizeUnit",
    "samplingProtocol": "http://rs.tdwg.org/dwc/terms/samplingProtocol",
}

META_OCCURRENCE_TERMS = {
    "eventID": "http://rs.tdwg.org/dwc/terms/eventID",
    "occurrenceID": "http://rs.tdwg.org/dwc/terms/occurrenceID",
    "scientificName": "http://rs.tdwg.org/dwc/terms/scientificName",
    "scientificNameID": "http://rs.tdwg.org/dwc/terms/scientificNameID",
    "taxonRank": "http://rs.tdwg.org/dwc/terms/taxonRank",
    "kingdom": "http://rs.tdwg.org/dwc/terms/kingdom",
    "basisOfRecord": "http://rs.tdwg.org/dwc/terms/basisOfRecord",
    "occurrenceStatus": "http://rs.tdwg.org/dwc/terms/occurrenceStatus",
    "verbatimIdentification": "http://rs.tdwg.org/dwc/terms/verbatimIdentification",
    "identifiedBy": "http://rs.tdwg.org/dwc/terms/identifiedBy",
    "identificationVerificationStatus": "http://rs.tdwg.org/dwc/terms/identificationVerificationStatus",
    "identificationReferences": "http://rs.tdwg.org/dwc/terms/identificationReferences",
    "associatedMedia": "http://rs.tdwg.org/dwc/terms/associatedMedia",
}

ATTRIBUTE_TEMPLATE_MAP = {
    "datasetName": {
        "attributeLabel": None,
        "attributeDefinition": "The name identifying the data set from which the record was derived.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/datasetName"),
    },
    "eventID": {
        "attributeLabel": "NA",
        "attributeDefinition": "An identifier for the set of information associated with an Event (something that occurs at a place and time). May be a global unique identifier or an identifier specific to the data set.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/eventID"),
    },
    "eventDate": {
        "attributeLabel": "NA",
        "attributeDefinition": "The date-time or interval during which an Event occurred. For occurrences, this is the date-time when the event was recorded. Not suitable for a time in a geological context.",
        "measurementScale": ("dateTime", "YYYY-MM-DDThh:mm:ss+00:00", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/eventDate"),
    },
    "geodeticDatum": {
        "attributeLabel": "NA",
        "attributeDefinition": "The ellipsoid, geodetic datum, or spatial reference system (SRS) upon which the geographic coordinates given in decimalLatitude and decimalLongitude as based.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/geodeticDatum"),
    },
    "countryCode": {
        "attributeLabel": "NA",
        "attributeDefinition": "The standard code for the country in which the Location occurs.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/countryCode"),
    },
    "institutionCode": {
        "attributeLabel": None,
        "attributeDefinition": "The name (or acronym) in use by the institution having custody of the object(s) or information referred to in the record.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/institutionCode"),
    },
    "decimalLatitude": {
        "attributeLabel": "NA",
        "attributeDefinition": "The geographic latitude (in decimal degrees, using the spatial reference system given in geodeticDatum) of the geographic center of a Location. Positive values are north of the Equator, negative values are south of it. Legal values lie between -90 and 90, inclusive.",
        "measurementScale": ("ratio", "degree", "real"),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/decimalLatitude"),
    },
    "decimalLongitude": {
        "attributeLabel": "NA",
        "attributeDefinition": "The geographic longitude (in decimal degrees, using the spatial reference system given in geodeticDatum) of the geographic center of a Location. Positive values are east of the Greenwich Meridian, negative values are west of it. Legal values lie between -180 and 180, inclusive.",
        "measurementScale": ("ratio", "degree", "real"),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/decimalLongitude"),
    },
    "minimumDepthInMeters": {
        "attributeLabel": "NA",
        "attributeDefinition": "The lesser depth of a range of depth below the local surface, in meters.",
        "measurementScale": ("interval", "meter", "real"),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/minimumDepthInMeters"),
    },
    "maximumDepthInMeters": {
        "attributeLabel": "NA",
        "attributeDefinition": "The greater depth of a range of depth below the local surface, in meters.",
        "measurementScale": ("interval", "meter", "real"),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/maximumDepthInMeters"),
    },
    "sampleSizeUnit": {
        "attributeLabel": None,
        "attributeDefinition": "The unit of measurement of the size (time duration, length, area, or volume) of a sample in a sampling dwc:Event.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/sampleSizeUnit"),
    },
    "sampleSizeValue": {
        "attributeLabel": None,
        "attributeDefinition": "A numeric value for a measurement of the size (time duration, length, area, or volume) of a sample in a sampling dwc:Event.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/sampleSizeValue"),
    },
    "samplingProtocol": {
        "attributeLabel": None,
        "attributeDefinition": "The names of, references to, or descriptions of the methods or protocols used during a dwc:Event.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/samplingProtocol"),
    },
    "basisOfRecord": {
        "attributeLabel": "NA",
        "attributeDefinition": "The specific nature of the data record.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/basisOfRecord"),
    },
    "occurrenceID": {
        "attributeLabel": "NA",
        "attributeDefinition": "An identifier for the Occurrence (as opposed to a particular digital record of the occurrence). In the absence of a persistent global unique identifier, construct one from a combination of identifiers in the record that will most closely make the occurrenceID globally unique.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/occurrenceID"),
    },
    "occurrenceStatus": {
        "attributeLabel": None,
        "attributeDefinition": "A statement about the presence or absence of a Taxon at a Location.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/occurrenceStatus"),
    },
    "taxonRank": {
        "attributeLabel": None,
        "attributeDefinition": "The taxonomic rank of the most specific name in the dwc:scientificName.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/taxonRank"),
    },
    "scientificName": {
        "attributeLabel": "NA",
        "attributeDefinition": "The full scientific name, with authorship and date information if known. When forming part of an Identification, this should be the name in lowest level taxonomic rank that can be determined. This term should not contain identification qualifications, which should instead be supplied in the IdentificationQualifier term.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/scientificName"),
    },
    "scientificNameID": {
        "attributeLabel": "NA",
        "attributeDefinition": "An identifier for the nomenclatural (not taxonomic) details of a scientific name.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/scientificNameID"),
    },
    "kingdom": {
        "attributeLabel": None,
        "attributeDefinition": "The full scientific name of the kingdom in which the taxon is classified.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/kingdom"),
    },
    "verbatimIdentification": {
        "attributeLabel": None,
        "attributeDefinition": "A string representing the taxonomic identification as it appeared in the original record.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/verbatimIdentification"),
    },
    "identifiedBy": {
        "attributeLabel": None,
        "attributeDefinition": "A list (concatenated and separated) of names of people, groups, or organizations who assigned the dwc:Taxon to the subject.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/identifiedBy"),
    },
    "identificationVerificationStatus": {
        "attributeLabel": None,
        "attributeDefinition": "A categorical indicator of the extent to which the taxonomic identification has been verified to be correct.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/identificationVerificationStatus"),
    },
    "identificationReferences": {
        "attributeLabel": None,
        "attributeDefinition": "A list (concatenated and separated) of references (publication, global unique identifier, URI) used in the dwc:Identification.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/identificationReferences"),
    },
    "associatedMedia": {
        "attributeLabel": None,
        "attributeDefinition": "A list (concatenated and separated) of identifiers (publication, global unique identifier, URI) of media associated with the dwc:Occurrence.",
        "measurementScale": ("nominal", "text", None),
        "annotation": ("http://ecoinformatics.org/oboe/oboe.1.2/oboe-core.owl#containsMeasurementsOfType", "http://rs.tdwg.org/dwc/terms/associatedMedia"),
    },
}

UNMATCHED_EVENT_FIELDS = [
    field for field in EVENT_FIELDS if field not in ATTRIBUTE_TEMPLATE_MAP
]
UNMATCHED_OCCURRENCE_FIELDS = [
    field for field in OCCURRENCE_FIELDS if field not in ATTRIBUTE_TEMPLATE_MAP
]

_APHIA_CLIENT = None
_APHIA_CLIENT_INITIALIZED = False
_TAXONOMY_CACHE: dict[str, Tuple[str, str]] = {}


@dataclass
class SeaBASSFile:
    header: dict[str, str]
    rows: list[dict[str, str]]


@dataclass
class AggregateMetadata:
    west_bounding_coordinate: Optional[float] = None
    east_bounding_coordinate: Optional[float] = None
    north_bounding_coordinate: Optional[float] = None
    south_bounding_coordinate: Optional[float] = None
    begin_date: Optional[str] = None
    end_date: Optional[str] = None


@dataclass
class MetadataTemplate:
    general: dict[str, str]
    creators: list[dict[str, str]]
    providers: list[dict[str, str]]
    contacts: list[dict[str, str]]
    associated_parties: list[dict[str, str]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Translate one SeaBASS IFCB .sb file or a directory of .sb files into first-pass Darwin Core CSV tables."
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="Path to a SeaBASS .sb file or a directory containing .sb files",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("out"),
        help="Directory for event.csv and occurrence.csv",
    )
    parser.add_argument("--country-code", default="US", help="Darwin Core countryCode")
    parser.add_argument(
        "--geodetic-datum", default="WGS84", help="Darwin Core geodeticDatum"
    )
    parser.add_argument(
        "--sampling-protocol",
        default="",
        help="Darwin Core samplingProtocol",
    )
    parser.add_argument(
        "--identification-references",
        default="",
        help="Darwin Core identificationReferences",
    )
    parser.add_argument(
        "--disable-taxonomy-lookup",
        action="store_true",
        help="Skip WoRMS lookups for taxonRank and kingdom",
    )
    parser.add_argument(
        "--metadata-template",
        type=Path,
        default=None,
        help="Path to the metadata template workbook used to generate eml.xml",
    )
    return parser.parse_args()


def strip_bracket_suffix(value: str) -> str:
    return value.split("[", 1)[0]


def parse_event_date(start_date: str, start_time: str) -> str:
    date_part = start_date.strip()
    time_part = strip_bracket_suffix(start_time).strip()
    parsed = datetime.strptime(f"{date_part} {time_part}", "%Y%m%d %H:%M:%S")
    return parsed.replace(tzinfo=timezone.utc).isoformat()


def parse_date_only(value: str) -> str:
    return datetime.strptime(value.strip(), "%Y%m%d").date().isoformat()


def parse_coordinate(value: str) -> Optional[float]:
    stripped = strip_bracket_suffix(value).strip()
    if not stripped or stripped == "NA":
        return None
    return float(stripped)


def clean_metadata_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_metadata_template(path: Path) -> MetadataTemplate:
    workbook = load_workbook(path, data_only=True)

    if "general" not in workbook.sheetnames:
        raise ValueError("Metadata template is missing required 'general' sheet")
    if "personnel" not in workbook.sheetnames:
        raise ValueError("Metadata template is missing required 'personnel' sheet")

    general_sheet = workbook["general"]
    personnel_sheet = workbook["personnel"]

    general: dict[str, str] = {}
    for row in general_sheet.iter_rows(min_row=2, values_only=True):
        key = clean_metadata_value(row[0] if len(row) > 0 else None)
        value = clean_metadata_value(row[1] if len(row) > 1 else None)
        if key:
            general[key] = value

    def load_people(sheet_name: str, worksheet) -> list[dict[str, str]]:
        headers = [clean_metadata_value(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            person = {
                header: clean_metadata_value(value)
                for header, value in zip(headers, row)
                if header
            }
            if any(person.values()):
                rows.append(person)
        if not rows:
            raise ValueError(
                f"Metadata template requires at least one entry in '{sheet_name}' sheet"
            )
        return rows

    personnel = load_people("personnel", personnel_sheet)

    creators = [
        person for person in personnel if person.get("role", "").strip().lower() == "creator"
    ]
    providers = [
        person for person in personnel if person.get("role", "").strip().lower() == "provider"
    ]
    contacts = [
        person for person in personnel if person.get("role", "").strip().lower() == "contact"
    ]
    associated_parties = [
        person
        for person in personnel
        if person.get("role", "").strip().lower() not in {"creator", "provider", "contact"}
    ]

    if not creators:
        raise ValueError(
            "Metadata template requires at least one personnel entry with role 'creator'"
        )
    if not providers:
        raise ValueError(
            "Metadata template requires at least one personnel entry with role 'provider'"
        )

    required_general_fields = [
        "title",
        "abstract",
        "geographicDescription",
        "identificationReferences",
        "publisherOrganization",
        "publisherEmail",
        "publisherURL",
    ]
    missing = [field for field in required_general_fields if not general.get(field)]
    if missing:
        raise ValueError(
            "Metadata template is missing required general values: "
            + ", ".join(missing)
        )

    return MetadataTemplate(
        general=general,
        creators=creators,
        providers=providers,
        contacts=contacts,
        associated_parties=associated_parties,
    )


def update_aggregate_metadata(
    aggregate: AggregateMetadata, sb_file: SeaBASSFile
) -> None:
    header = sb_file.header

    west = parse_coordinate(header.get("west_longitude", ""))
    east = parse_coordinate(header.get("east_longitude", ""))
    north = parse_coordinate(header.get("north_latitude", ""))
    south = parse_coordinate(header.get("south_latitude", ""))

    if west is not None:
        if aggregate.west_bounding_coordinate is None:
            aggregate.west_bounding_coordinate = west
        else:
            aggregate.west_bounding_coordinate = min(
                aggregate.west_bounding_coordinate, west
            )

    if east is not None:
        if aggregate.east_bounding_coordinate is None:
            aggregate.east_bounding_coordinate = east
        else:
            aggregate.east_bounding_coordinate = max(
                aggregate.east_bounding_coordinate, east
            )

    if north is not None:
        if aggregate.north_bounding_coordinate is None:
            aggregate.north_bounding_coordinate = north
        else:
            aggregate.north_bounding_coordinate = max(
                aggregate.north_bounding_coordinate, north
            )

    if south is not None:
        if aggregate.south_bounding_coordinate is None:
            aggregate.south_bounding_coordinate = south
        else:
            aggregate.south_bounding_coordinate = min(
                aggregate.south_bounding_coordinate, south
            )

    start_date = header.get("start_date", "")
    if start_date:
        event_start = parse_date_only(start_date)
        if aggregate.begin_date is None or event_start < aggregate.begin_date:
            aggregate.begin_date = event_start

    end_date = header.get("end_date", "")
    if end_date:
        event_end = parse_date_only(end_date)
        if aggregate.end_date is None or event_end > aggregate.end_date:
            aggregate.end_date = event_end


def extract_taxon_token(scientific_name_id: str) -> str:
    if not scientific_name_id:
        return "unknown"
    return scientific_name_id.rsplit(":", 1)[-1]


def extract_aphia_id(scientific_name_id: str) -> Optional[str]:
    prefix = "urn:lsid:marinespecies.org:taxname:"
    if not scientific_name_id.startswith(prefix):
        return None
    aphia_id = scientific_name_id[len(prefix) :].strip()
    return aphia_id or None


def fetch_taxonomy(aphia_id: str, disabled: bool) -> Tuple[str, str]:
    global _APHIA_CLIENT
    global _APHIA_CLIENT_INITIALIZED
    global _TAXONOMY_CACHE

    if disabled:
        return "", ""

    if aphia_id in _TAXONOMY_CACHE:
        return _TAXONOMY_CACHE[aphia_id]

    try:
        from suds.client import Client  # noqa: F401
        from third_party.worms import Aphia
    except ImportError:
        print(
            "Warning: worms.py dependency 'suds' is not installed; leaving taxonRank and kingdom blank.",
            file=sys.stderr,
        )
        _TAXONOMY_CACHE[aphia_id] = ("", "")
        return _TAXONOMY_CACHE[aphia_id]

    try:
        if not _APHIA_CLIENT_INITIALIZED:
            _APHIA_CLIENT = Aphia()
            _APHIA_CLIENT_INITIALIZED = True

        if _APHIA_CLIENT is None or getattr(_APHIA_CLIENT, "client", None) is None:
            _TAXONOMY_CACHE[aphia_id] = ("", "")
            return _TAXONOMY_CACHE[aphia_id]
        record = _APHIA_CLIENT.get_aphia_record_by_id(int(aphia_id))
    except Exception as exc:  # pragma: no cover - network/runtime dependent
        print(
            f"Warning: worms.py lookup failed for AphiaID {aphia_id}: {exc}",
            file=sys.stderr,
        )
        _TAXONOMY_CACHE[aphia_id] = ("", "")
        return _TAXONOMY_CACHE[aphia_id]

    if not record:
        _TAXONOMY_CACHE[aphia_id] = ("", "")
        return _TAXONOMY_CACHE[aphia_id]

    _TAXONOMY_CACHE[aphia_id] = (
        getattr(record, "rank", "") or "",
        getattr(record, "kingdom", "") or "",
    )
    return _TAXONOMY_CACHE[aphia_id]


def parse_sb_file(path: Path) -> SeaBASSFile:
    header: dict[str, str] = {}
    rows: list[dict[str, str]] = []
    in_header = False
    field_names: list[str] = []

    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if line == "/begin_header":
            in_header = True
            continue
        if line == "/end_header":
            in_header = False
            continue

        if in_header:
            if raw_line.startswith("/") and "=" in raw_line:
                key, value = raw_line[1:].split("=", 1)
                header[key] = value
                if key == "fields":
                    field_names = value.split(",")
            continue

        if not line or line.startswith("!"):
            continue

        if not field_names:
            raise ValueError("No /fields definition found before data rows")

        # SeaBASS body rows are interpreted against the /fields declaration,
        # so downstream mapping is based on field names rather than fixed order.
        values = next(csv.reader([raw_line]))
        if len(values) != len(field_names):
            raise ValueError(
                f"Field count mismatch in {path.name}: expected {len(field_names)}, got {len(values)}"
            )
        rows.append(dict(zip(field_names, values)))

    return SeaBASSFile(header=header, rows=rows)


def collect_input_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.glob("*.sb"))
    raise ValueError(f"Input path does not exist: {input_path}")


def build_event_row(sb_file: SeaBASSFile, args: argparse.Namespace) -> dict[str, str]:
    header = sb_file.header
    measurement_depth = header.get("measurement_depth", "")
    return {
        "datasetName": header.get("cruise", ""),
        "eventID": header.get("eventID", ""),
        "eventDate": parse_event_date(
            header.get("start_date", ""), header.get("start_time", "")
        ),
        "institutionCode": header.get("affiliations", ""),
        "decimalLongitude": strip_bracket_suffix(header.get("east_longitude", "")).strip(),
        "decimalLatitude": strip_bracket_suffix(header.get("north_latitude", "")).strip(),
        "countryCode": args.country_code,
        "geodeticDatum": args.geodetic_datum,
        "minimumDepthInMeters": measurement_depth,
        "maximumDepthInMeters": measurement_depth,
        "sampleSizeValue": header.get("volume_sampled_ml", ""),
        "sampleSizeUnit": "milliliter",
        "samplingProtocol": "",
    }


def build_occurrence_rows(
    sb_file: SeaBASSFile,
    args: argparse.Namespace,
    metadata_template: Optional[MetadataTemplate] = None,
) -> list[dict[str, str]]:
    event_id = sb_file.header.get("eventID", "")
    taxon_counts: defaultdict[str, int] = defaultdict(int)
    occurrence_rows: list[dict[str, str]] = []
    identification_references = args.identification_references
    if metadata_template is not None:
        identification_references = metadata_template.general.get(
            "identificationReferences", identification_references
        )

    for row in sb_file.rows:
        scientific_name_id = row.get("scientificNameID_automated", "")
        aphia_id = extract_aphia_id(scientific_name_id)
        if not aphia_id:
            continue

        taxon_token = extract_taxon_token(scientific_name_id)
        taxon_counts[taxon_token] += 1
        occurrence_id = f"{event_id}_{taxon_token}_{taxon_counts[taxon_token]}"
        taxon_rank, kingdom = fetch_taxonomy(aphia_id, args.disable_taxonomy_lookup)
        occurrence_rows.append(
            {
                "eventID": event_id,
                "occurrenceID": occurrence_id,
                "scientificName": row.get("scientificName_automated", ""),
                "scientificNameID": scientific_name_id,
                "taxonRank": taxon_rank,
                "kingdom": kingdom,
                "basisOfRecord": "MachineObservation",
                "occurrenceStatus": "present",
                "verbatimIdentification": row.get(
                    "data_provider_category_automated", ""
                ),
                "identifiedBy": "",
                "identificationVerificationStatus": "PredictedByMachine",
                "identificationReferences": identification_references,
                "associatedMedia": row.get("associatedMedia", ""),
            }
        )

    return occurrence_rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def indent_xml(element: ET.Element, level: int = 0) -> None:
    indent = "\n" + level * "  "
    if len(element):
        if not element.text or not element.text.strip():
            element.text = indent + "  "
        for child in element:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = indent
    elif level and (not element.tail or not element.tail.strip()):
        element.tail = indent


def build_meta_xml() -> ET.Element:
    archive = ET.Element(
        "archive",
        {
            "xmlns": "http://rs.tdwg.org/dwc/text/",
            "metadata": "eml.xml",
        },
    )

    core = ET.SubElement(
        archive,
        "core",
        {
            "encoding": "UTF-8",
            "fieldsTerminatedBy": ",",
            "linesTerminatedBy": "\\n",
            "fieldsEnclosedBy": "",
            "ignoreHeaderLines": "1",
            "rowType": "http://rs.tdwg.org/dwc/terms/Event",
        },
    )
    core_files = ET.SubElement(core, "files")
    ET.SubElement(core_files, "location").text = "event.csv"
    ET.SubElement(core, "id", {"index": str(META_EVENT_ID_INDEX)})

    for index, field_name in enumerate(EVENT_FIELDS):
        if index == META_EVENT_ID_INDEX:
            continue
        ET.SubElement(
            core,
            "field",
            {
                "index": str(index),
                "term": META_EVENT_TERMS[field_name],
            },
        )

    occurrence = ET.SubElement(
        archive,
        "extension",
        {
            "encoding": "UTF-8",
            "fieldsTerminatedBy": ",",
            "linesTerminatedBy": "\\n",
            "fieldsEnclosedBy": "",
            "ignoreHeaderLines": "1",
            "rowType": "http://rs.tdwg.org/dwc/terms/Occurrence",
        },
    )
    occurrence_files = ET.SubElement(occurrence, "files")
    ET.SubElement(occurrence_files, "location").text = "occurrence.csv"
    ET.SubElement(
        occurrence,
        "coreid",
        {"index": str(META_OCCURRENCE_CORE_ID_INDEX)},
    )

    for index, field_name in enumerate(OCCURRENCE_FIELDS):
        if index == META_OCCURRENCE_CORE_ID_INDEX:
            continue
        ET.SubElement(
            occurrence,
            "field",
            {
                "index": str(index),
                "term": META_OCCURRENCE_TERMS[field_name],
            },
        )

    indent_xml(archive)
    return archive


def write_meta_xml(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tree = ET.ElementTree(build_meta_xml())
    if hasattr(ET, "indent"):
        ET.indent(tree, space="  ")
    tree.write(path, encoding="UTF-8", xml_declaration=False)


def append_text_element(parent: ET.Element, tag: str, text: str) -> ET.Element:
    element = ET.SubElement(parent, tag)
    element.text = text
    return element


def append_measurement_scale(
    parent: ET.Element,
    scale_type: str,
    value: str,
    number_type: Optional[str],
    default_definition: Optional[str] = None,
) -> None:
    measurement_scale = ET.SubElement(parent, "measurementScale")
    if scale_type == "nominal":
        nominal = ET.SubElement(measurement_scale, "nominal")
        non_numeric = ET.SubElement(nominal, "nonNumericDomain")
        text_domain = ET.SubElement(non_numeric, "textDomain")
        append_text_element(text_domain, "definition", default_definition or value)
    elif scale_type == "dateTime":
        date_time = ET.SubElement(measurement_scale, "dateTime")
        append_text_element(date_time, "formatString", value)
    elif scale_type in {"ratio", "interval"}:
        scale = ET.SubElement(measurement_scale, scale_type)
        unit = ET.SubElement(scale, "unit")
        append_text_element(unit, "standardUnit", value)
        numeric_domain = ET.SubElement(scale, "numericDomain")
        append_text_element(numeric_domain, "numberType", number_type or "real")
    else:
        nominal = ET.SubElement(measurement_scale, "nominal")
        non_numeric = ET.SubElement(nominal, "nonNumericDomain")
        text_domain = ET.SubElement(non_numeric, "textDomain")
        append_text_element(text_domain, "definition", value)


def append_missing_value_code(parent: ET.Element) -> None:
    missing = ET.SubElement(parent, "missingValueCode")
    append_text_element(missing, "code", "NA")
    append_text_element(missing, "codeExplanation", "Missing value")


def build_attribute_element(
    parent: ET.Element, entity_id: str, field_name: str, template: dict[str, Any]
) -> None:
    attribute = ET.SubElement(parent, "attribute", {"id": f"{entity_id}/{field_name}"})
    append_text_element(attribute, "attributeName", field_name)
    if template.get("attributeLabel") is not None:
        append_text_element(attribute, "attributeLabel", template["attributeLabel"])
    append_text_element(attribute, "attributeDefinition", template["attributeDefinition"])
    scale_type, value, number_type = template["measurementScale"]
    append_measurement_scale(
        attribute,
        scale_type,
        value,
        number_type,
        default_definition=template["attributeDefinition"],
    )
    append_missing_value_code(attribute)
    annotation = template.get("annotation")
    if annotation is not None:
        annotation_element = ET.SubElement(attribute, "annotation")
        append_text_element(
            annotation_element, "propertyURI", annotation[0]
        ).set("label", "contains measurements of type")
        append_text_element(annotation_element, "valueURI", annotation[1]).set(
            "label", field_name
        )


def build_fallback_attribute_template(field_name: str) -> dict[str, Any]:
    return {
        "attributeLabel": None,
        "attributeDefinition": f"No sample EML attribute template was available for {field_name}.",
        "measurementScale": ("nominal", f"No sample EML attribute template was available for {field_name}.", None),
        "annotation": None,
    }


def add_data_table(
    dataset: ET.Element,
    object_name: str,
    entity_name: str,
    entity_description: str,
    file_path: Path,
    field_names: list[str],
    record_count: int,
    annotation_value_uri: Optional[str] = None,
    annotation_value_label: Optional[str] = None,
) -> None:
    data_table = ET.SubElement(dataset, "dataTable", {"id": object_name})
    append_text_element(data_table, "entityName", entity_name)
    append_text_element(data_table, "entityDescription", entity_description)

    physical = ET.SubElement(data_table, "physical")
    append_text_element(physical, "objectName", object_name)
    size = append_text_element(physical, "size", str(file_path.stat().st_size))
    size.set("unit", "bytes")
    data_format = ET.SubElement(physical, "dataFormat")
    external = ET.SubElement(data_format, "externallyDefinedFormat")
    append_text_element(external, "formatName", "text/csv")

    if annotation_value_uri and annotation_value_label:
        annotation = ET.SubElement(data_table, "annotation")
        append_text_element(
            annotation, "propertyURI", "http://purl.obolibrary.org/obo/IAO_0000136"
        ).set("label", "is about")
        append_text_element(annotation, "valueURI", annotation_value_uri).set(
            "label", annotation_value_label
        )

    attribute_list = ET.SubElement(data_table, "attributeList")
    for field_name in field_names:
        template = ATTRIBUTE_TEMPLATE_MAP.get(field_name)
        if template is None:
            template = build_fallback_attribute_template(field_name)
        build_attribute_element(attribute_list, object_name, field_name, template)

    append_text_element(data_table, "numberOfRecords", str(record_count))


def build_party_element(parent: ET.Element, tag: str, person: dict[str, str]) -> None:
    party = ET.SubElement(parent, tag)
    individual_name = ET.SubElement(party, "individualName")
    given_name = person.get("givenName", "")
    middle_initial = person.get("middleInitial", "")
    if middle_initial:
        given_name = f"{given_name} {middle_initial}".strip()
    append_text_element(individual_name, "givenName", given_name)
    append_text_element(individual_name, "surName", person.get("surName", ""))

    if person.get("organizationName"):
        append_text_element(party, "organizationName", person["organizationName"])

    position_name = person.get("positionName") or person.get("role", "")
    if position_name:
        append_text_element(party, "positionName", position_name)

    if person.get("electronicMailAddress"):
        append_text_element(
            party,
            "electronicMailAddress",
            person["electronicMailAddress"],
        )

    if tag in {"associatedParty", "personnel"} and person.get("role"):
        append_text_element(party, "role", person["role"])


def maybe_add_taxonomic_coverage(coverage: ET.Element) -> None:
    """Placeholder hook for future EML taxonomicCoverage generation."""
    # Intentionally left as a no-op until taxonomicCoverage requirements
    # and source data rules are finalized.
    _ = coverage


def build_eml_xml(
    metadata: MetadataTemplate,
    aggregate: AggregateMetadata,
    archive_name: str,
    run_date: date,
    output_dir: Path,
    event_record_count: int,
    occurrence_record_count: int,
) -> ET.ElementTree:
    attributes = {
        "xmlns:eml": "eml://ecoinformatics.org/eml-2.1.1",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "xml:lang": "eng",
        "packageId": archive_name.replace(".zip", ""),
        "xsi:schemaLocation": (
            "eml://ecoinformatics.org/eml-2.1.1 "
            "http://rs.gbif.org/schema/eml-gbif-profile/1.1/eml.xsd"
        ),
        "scope": "system",
        "system": "http://gbif.org",
    }
    root = ET.Element("eml:eml", attributes)
    dataset = ET.SubElement(root, "dataset")

    append_text_element(dataset, "title", metadata.general["title"])
    for person in metadata.creators:
        build_party_element(dataset, "creator", person)
    for person in metadata.providers:
        build_party_element(dataset, "metadataProvider", person)
    for person in metadata.associated_parties:
        build_party_element(dataset, "associatedParty", person)

    append_text_element(dataset, "pubDate", run_date.isoformat())
    append_text_element(dataset, "language", "eng")

    abstract = ET.SubElement(dataset, "abstract")
    append_text_element(abstract, "para", metadata.general["abstract"])

    keyword_set = ET.SubElement(dataset, "keywordSet")
    append_text_element(keyword_set, "keyword", "samplingEvent")
    append_text_element(
        keyword_set,
        "keywordThesaurus",
        "GBIF Dataset Type Vocabulary: http://rs.gbif.org/vocabulary/gbif/dataset_type.xml",
    )

    additional_info = ET.SubElement(dataset, "additionalInfo")
    append_text_element(additional_info, "para", "marine, harvested by OBIS")

    rights = ET.SubElement(dataset, "intellectualRights")
    rights_para = ET.SubElement(rights, "para")
    rights_para.text = "This work is licensed under a "
    rights_link = ET.SubElement(
        rights_para,
        "ulink",
        {"url": "http://creativecommons.org/licenses/by/4.0/legalcode"},
    )
    append_text_element(
        rights_link,
        "citetitle",
        "Creative Commons Attribution (CC-BY) 4.0 License",
    )
    rights_link.tail = "."

    coverage = ET.SubElement(dataset, "coverage")
    geographic_coverage = ET.SubElement(coverage, "geographicCoverage")
    append_text_element(
        geographic_coverage,
        "geographicDescription",
        metadata.general["geographicDescription"],
    )
    bounding_coordinates = ET.SubElement(geographic_coverage, "boundingCoordinates")
    append_text_element(
        bounding_coordinates,
        "westBoundingCoordinate",
        str(aggregate.west_bounding_coordinate or ""),
    )
    append_text_element(
        bounding_coordinates,
        "eastBoundingCoordinate",
        str(aggregate.east_bounding_coordinate or ""),
    )
    append_text_element(
        bounding_coordinates,
        "northBoundingCoordinate",
        str(aggregate.north_bounding_coordinate or ""),
    )
    append_text_element(
        bounding_coordinates,
        "southBoundingCoordinate",
        str(aggregate.south_bounding_coordinate or ""),
    )

    maybe_add_taxonomic_coverage(coverage)

    temporal_coverage = ET.SubElement(coverage, "temporalCoverage")
    range_of_dates = ET.SubElement(temporal_coverage, "rangeOfDates")
    begin_date = ET.SubElement(range_of_dates, "beginDate")
    append_text_element(begin_date, "calendarDate", aggregate.begin_date or "")
    end_date = ET.SubElement(range_of_dates, "endDate")
    append_text_element(end_date, "calendarDate", aggregate.end_date or "")

    maintenance = ET.SubElement(dataset, "maintenance")
    maintenance_description = ET.SubElement(maintenance, "description")
    ET.SubElement(maintenance_description, "para")
    append_text_element(
        maintenance, "maintenanceUpdateFrequency", "irregular"
    )

    for person in metadata.contacts:
        build_party_element(dataset, "contact", person)

    project = ET.SubElement(dataset, "project")
    append_text_element(project, "title", metadata.general["title"])
    project_person = (
        metadata.creators[0]
        if metadata.creators
        else metadata.contacts[0]
        if metadata.contacts
        else metadata.providers[0]
    )
    build_party_element(project, "personnel", project_person)

    return ET.ElementTree(root)


def write_eml_xml(
    path: Path,
    metadata: MetadataTemplate,
    aggregate: AggregateMetadata,
    archive_name: str,
    run_date: date,
    output_dir: Path,
    event_record_count: int,
    occurrence_record_count: int,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tree = build_eml_xml(
        metadata,
        aggregate,
        archive_name,
        run_date,
        output_dir,
        event_record_count,
        occurrence_record_count,
    )
    if hasattr(ET, "indent"):
        ET.indent(tree, space="  ")
    tree.write(path, encoding="UTF-8", xml_declaration=True)


def build_archive_name(event_rows: list[dict[str, str]]) -> str:
    dataset_names = sorted({row.get("datasetName", "").strip() for row in event_rows if row.get("datasetName", "").strip()})
    if len(dataset_names) == 1:
        return f"{dataset_names[0]}_dwc.zip"
    if len(dataset_names) > 1:
        return "combined_dwc.zip"
    return "dwc.zip"


def write_archive(output_dir: Path, archive_name: str) -> None:
    archive_path = output_dir / archive_name
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(output_dir / "event.csv", arcname="event.csv")
        archive.write(output_dir / "occurrence.csv", arcname="occurrence.csv")
        meta_xml = output_dir / "meta.xml"
        if meta_xml.exists():
            archive.write(meta_xml, arcname="meta.xml")
        eml_xml = output_dir / "eml.xml"
        if eml_xml.exists():
            archive.write(eml_xml, arcname="eml.xml")


def main() -> None:
    args = parse_args()
    input_files = collect_input_files(args.input_path)
    if not input_files:
        raise ValueError(f"No .sb files found in {args.input_path}")

    metadata_template = (
        load_metadata_template(args.metadata_template)
        if args.metadata_template is not None
        else None
    )

    event_rows: list[dict[str, str]] = []
    occurrence_rows: list[dict[str, str]] = []
    aggregate_metadata = AggregateMetadata()

    for input_file in input_files:
        sb_file = parse_sb_file(input_file)
        update_aggregate_metadata(aggregate_metadata, sb_file)
        event_rows.append(build_event_row(sb_file, args))
        occurrence_rows.extend(build_occurrence_rows(sb_file, args, metadata_template))
        print(
            f"Processed {len(event_rows)} of {len(input_files)} files: {input_file.name}",
            file=sys.stderr,
        )

    write_csv(args.output_dir / "event.csv", EVENT_FIELDS, event_rows)
    write_csv(args.output_dir / "occurrence.csv", OCCURRENCE_FIELDS, occurrence_rows)
    write_meta_xml(args.output_dir / "meta.xml")
    archive_name = build_archive_name(event_rows)
    if metadata_template is not None:
        write_eml_xml(
            args.output_dir / "eml.xml",
            metadata_template,
            aggregate_metadata,
            archive_name,
            date.today(),
            args.output_dir,
            len(event_rows),
            len(occurrence_rows),
        )
    write_archive(args.output_dir, archive_name)


if __name__ == "__main__":
    main()
